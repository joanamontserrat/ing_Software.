import tkinter as tk
from tkinter import messagebox, filedialog
from datos.datos_animales import animales, adopciones
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class Reportes:
    def __init__(self, master):
        tk.Label(master, text="Reportes / Exportar Datos 🖨️", font=("Arial", 18), bg="#F1BB81").pack(pady=20)

        frame_botones = tk.Frame(master, bg="#F1BB81")
        frame_botones.pack(pady=20)

        botones = [
            ("Exportar Animales en Adopción (Excel)", self.exportar_adopcion),
            ("Exportar Animales Adoptados (Excel)", self.exportar_adoptados),
            ("Exportar Animales en Recuperación (Excel)", self.exportar_recuperacion),
            ("Exportar Lista General (Excel)", self.exportar_general),
        ]

        for texto, comando in botones:
            tk.Button(frame_botones, text=texto, command=comando, bg="#F2854B", width=40).pack(pady=5)

    def seleccionar_archivo_guardar(self, nombre_por_defecto):
        return filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivo Excel", "*.xlsx")],
            initialfile=nombre_por_defecto,
            title="Guardar reporte como"
        )

    def exportar_adopcion(self):
        en_adopcion = [a for a in animales if a.get("estado_actual") == "En Adopción"]
        ruta = self.seleccionar_archivo_guardar("animales_en_adopcion.xlsx")
        if ruta:
            self.exportar_a_excel(en_adopcion, ruta, columnas=["nombre", "especie", "edad", "estado_actual", "salud"])

    def exportar_recuperacion(self):
        en_recuperacion = [
            a for a in animales
            if a.get("salud", "").lower() in ["grave", "regular"]
        ]
        ruta = self.seleccionar_archivo_guardar("animales_en_recuperacion.xlsx")
        if ruta:
            columnas = ["nombre", "especie", "raza", "edad", "salud"]
            self.exportar_a_excel(en_recuperacion, ruta, columnas=columnas)


    def exportar_adoptados(self):
        ruta = self.seleccionar_archivo_guardar("animales_adoptados.xlsx")
        if ruta:
            try:
                columnas = ["animal", "especie", "edad", "salud", "adoptante", "correo", "telefono", "fecha"]
                df = pd.DataFrame(adopciones)
                df = df[[col for col in columnas if col in df.columns]]
                df.to_excel(ruta, index=False)
                self.formatear_excel(ruta)
                messagebox.showinfo("Éxito", f"Archivo '{ruta}' generado exitosamente.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo exportar el archivo: {e}")

    def exportar_general(self):
        ruta = self.seleccionar_archivo_guardar("lista_general_animales.xlsx")
        if ruta:
            columnas = ["nombre", "especie", "raza", "edad", "salud", "estado_actual"]
            self.exportar_a_excel(animales, ruta, columnas=columnas)

    def exportar_a_excel(self, lista, ruta, columnas):
        try:
            df = pd.DataFrame(lista)
            df = df[[col for col in columnas if col in df.columns]]
            df.to_excel(ruta, index=False)
            self.formatear_excel(ruta)
            messagebox.showinfo("Éxito", f"Archivo '{ruta}' generado exitosamente.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar el archivo: {e}")

    def formatear_excel(self, ruta):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(ruta)
            ws = wb.active

            # Formato: encabezados en negrita y centrados
            for col in range(1, ws.max_column + 1):
                celda = ws.cell(row=1, column=col)
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal="center", vertical="center")
                # Ajustar ancho de columnas automáticamente
                letra_col = get_column_letter(col)
                max_long = max(len(str(ws.cell(row=row, column=col).value or "")) for row in range(1, ws.max_row + 1))
                ws.column_dimensions[letra_col].width = max(15, min(40, max_long + 2))

            wb.save(ruta)
        except Exception as e:
            messagebox.showwarning("Formato parcial", f"El archivo fue generado pero no se pudo aplicar formato: {e}")
